#!/usr/bin/env python3
import csv
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    base_dir = Path(__file__).resolve().parent.parent
    xlsx_path = base_dir / "PAYSCALE23.xlsx"
    csv_path = base_dir / "PAYSCALE23.csv"

    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
        ):
            writer.writerow(["" if value is None else value for value in row])

    print(csv_path)


if __name__ == "__main__":
    main()
